"""SMS and guest management API endpoints."""
import io
import csv
import logging
from datetime import datetime, date
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Request, Header
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from core.database import get_db
from core.auth import get_current_user
from core.config import IS_PRODUCTION
from models.wedding import Wedding
from models.user import User
from models.sms import Guest, SMSTemplate, ScheduledMessage, MessageLog
from services.sms.twilio_service import twilio_service
from services.sms.template_service import template_service, DEFAULT_TEMPLATES

logger = logging.getLogger(__name__)

router = APIRouter()


# --- Pydantic Schemas ---

class GuestCreate(BaseModel):
    name: str
    phone_number: str
    email: Optional[str] = None
    group_name: Optional[str] = None
    rsvp_status: str = "pending"


class GuestUpdate(BaseModel):
    name: Optional[str] = None
    phone_number: Optional[str] = None
    email: Optional[str] = None
    group_name: Optional[str] = None
    rsvp_status: Optional[str] = None
    sms_consent: Optional[bool] = None


class GuestResponse(BaseModel):
    id: str
    name: str
    phone_number: str
    email: Optional[str]
    group_name: Optional[str]
    rsvp_status: str
    sms_consent: bool
    opted_out: bool
    created_at: datetime

    class Config:
        from_attributes = True


class TemplateCreate(BaseModel):
    name: str
    content: str
    category: str = "custom"


class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    content: Optional[str] = None
    category: Optional[str] = None


class TemplateResponse(BaseModel):
    id: str
    name: str
    content: str
    category: str
    is_default: bool
    created_at: datetime

    class Config:
        from_attributes = True


class SMSBlastRequest(BaseModel):
    message: str
    recipient_type: str = "all"  # all, group, individual
    recipient_filter: Optional[dict] = None  # {"group": "..."} or {"guest_ids": [...]}


class ScheduleMessageRequest(BaseModel):
    name: str
    message: str
    recipient_type: str = "all"
    recipient_filter: Optional[dict] = None
    schedule_type: str = "fixed"  # fixed or relative
    scheduled_at: Optional[datetime] = None  # For fixed
    relative_to: Optional[str] = None  # wedding_date, rsvp_deadline
    relative_days: Optional[int] = None  # -7 = 7 days before


class ScheduledMessageResponse(BaseModel):
    id: str
    name: str
    message_content: str
    recipient_type: str
    schedule_type: str
    scheduled_at: Optional[datetime]
    relative_to: Optional[str]
    relative_days: Optional[int]
    status: str
    sent_count: int
    failed_count: int
    total_recipients: int
    created_at: datetime

    class Config:
        from_attributes = True


class MessageLogResponse(BaseModel):
    id: str
    guest_id: str
    phone_number: str
    message_content: str
    status: str
    error_code: Optional[str]
    error_message: Optional[str]
    sent_at: Optional[datetime]
    delivered_at: Optional[datetime]

    class Config:
        from_attributes = True


# --- Helper Functions ---

async def get_wedding_for_user(db: AsyncSession, user: User) -> Wedding:
    """Get wedding for authenticated user."""
    if not user.wedding_id:
        raise HTTPException(status_code=404, detail="No wedding found for user")

    result = await db.execute(
        select(Wedding)
        .options(selectinload(Wedding.guests))
        .where(Wedding.id == user.wedding_id)
    )
    wedding = result.scalar_one_or_none()
    if not wedding:
        raise HTTPException(status_code=404, detail="Wedding not found")
    return wedding


async def get_eligible_recipients(
    db: AsyncSession,
    wedding_id: str,
    recipient_type: str,
    recipient_filter: Optional[dict]
) -> List[Guest]:
    """Get list of guests eligible for SMS (consented, not opted out)."""
    query = select(Guest).where(
        Guest.wedding_id == wedding_id,
        Guest.sms_consent == True,
        Guest.opted_out == False
    )

    if recipient_type == "group" and recipient_filter and "group" in recipient_filter:
        query = query.where(Guest.group_name == recipient_filter["group"])
    elif recipient_type == "individual" and recipient_filter and "guest_ids" in recipient_filter:
        query = query.where(Guest.id.in_(recipient_filter["guest_ids"]))

    result = await db.execute(query)
    return list(result.scalars().all())


# --- Guest Endpoints ---

@router.get("/{wedding_id}/guests", response_model=List[GuestResponse])
async def list_guests(
    wedding_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all guests for a wedding."""
    # Verify user owns this wedding
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(Guest).where(Guest.wedding_id == wedding_id).order_by(Guest.name)
    )
    return result.scalars().all()


@router.post("/{wedding_id}/guests", response_model=dict)
async def create_guest(
    wedding_id: str,
    data: GuestCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Add a single guest."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Validate phone number
    is_valid, formatted, country = twilio_service.validate_phone_number(data.phone_number)
    if not is_valid:
        raise HTTPException(status_code=400, detail=f"Invalid phone number: {formatted}")

    # Check for duplicate phone
    result = await db.execute(
        select(Guest).where(
            Guest.wedding_id == wedding_id,
            Guest.phone_number == formatted
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Guest with this phone number already exists")

    guest = Guest(
        wedding_id=wedding_id,
        name=data.name,
        phone_number=formatted,
        email=data.email,
        group_name=data.group_name,
        rsvp_status=data.rsvp_status
    )
    db.add(guest)
    await db.commit()
    await db.refresh(guest)

    return {"id": guest.id, "message": "Guest added successfully"}


@router.post("/{wedding_id}/guests/upload", response_model=dict)
async def upload_guests(
    wedding_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload guests from CSV or Excel file."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Validate file size (max 5MB)
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is 5MB, got {len(content) / 1024 / 1024:.1f}MB"
        )

    filename = file.filename or ""

    # Validate file extension
    if not filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload a CSV or Excel file (.csv, .xlsx, .xls)"
        )

    guests_added = 0
    guests_skipped = 0
    errors = []

    try:
        if filename.endswith(".csv"):
            # Parse CSV
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            # Parse Excel
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise HTTPException(status_code=400, detail="File must be CSV or Excel (.xlsx)")

        # Process rows
        for i, row in enumerate(rows, start=2):
            name = row.get("name") or row.get("Name") or row.get("Guest Name")
            phone = row.get("phone") or row.get("Phone") or row.get("phone_number") or row.get("Phone Number")
            email = row.get("email") or row.get("Email")
            group = row.get("group") or row.get("Group") or row.get("group_name")

            if not name or not phone:
                errors.append(f"Row {i}: Missing name or phone")
                guests_skipped += 1
                continue

            # Validate phone
            is_valid, formatted, _ = twilio_service.validate_phone_number(str(phone))
            if not is_valid:
                errors.append(f"Row {i}: Invalid phone number '{phone}'")
                guests_skipped += 1
                continue

            # Check for duplicate
            result = await db.execute(
                select(Guest).where(
                    Guest.wedding_id == wedding_id,
                    Guest.phone_number == formatted
                )
            )
            if result.scalar_one_or_none():
                guests_skipped += 1
                continue

            guest = Guest(
                wedding_id=wedding_id,
                name=str(name),
                phone_number=formatted,
                email=str(email) if email else None,
                group_name=str(group) if group else None
            )
            db.add(guest)
            guests_added += 1

        await db.commit()

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

    return {
        "message": f"Added {guests_added} guests, skipped {guests_skipped}",
        "added": guests_added,
        "skipped": guests_skipped,
        "errors": errors[:10]  # Return first 10 errors only
    }


@router.patch("/{wedding_id}/guests/{guest_id}", response_model=dict)
async def update_guest(
    wedding_id: str,
    guest_id: str,
    data: GuestUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a guest."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding_id)
    )
    guest = result.scalar_one_or_none()
    if not guest:
        raise HTTPException(status_code=404, detail="Guest not found")

    update_data = data.model_dump(exclude_unset=True)

    # Validate phone if being updated
    if "phone_number" in update_data:
        is_valid, formatted, _ = twilio_service.validate_phone_number(update_data["phone_number"])
        if not is_valid:
            raise HTTPException(status_code=400, detail=f"Invalid phone number: {formatted}")
        update_data["phone_number"] = formatted

    for field, value in update_data.items():
        setattr(guest, field, value)

    await db.commit()
    return {"id": guest_id, "message": "Guest updated"}


@router.delete("/{wedding_id}/guests/{guest_id}", response_model=dict)
async def delete_guest(
    wedding_id: str,
    guest_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a guest."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(Guest).where(Guest.id == guest_id, Guest.wedding_id == wedding_id)
    )
    guest = result.scalar_one_or_none()
    if not guest:
        raise HTTPException(status_code=404, detail="Guest not found")

    await db.delete(guest)
    await db.commit()
    return {"message": "Guest deleted"}


# --- Template Endpoints ---

@router.get("/{wedding_id}/templates", response_model=List[TemplateResponse])
async def list_templates(
    wedding_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all templates for a wedding, seeding defaults if none exist."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get templates for this wedding
    result = await db.execute(
        select(SMSTemplate).where(SMSTemplate.wedding_id == wedding_id)
    )
    templates = list(result.scalars().all())

    # If no templates exist, seed the default templates for this wedding
    if not templates:
        default_templates = [
            {
                "name": "Welcome Message",
                "category": "welcome",
                "content": (
                    "Hi {{guest_name}}! {{partner1}} & {{partner2}} are getting married! "
                    "Get all the details and ask questions here: {{chat_link}}"
                ),
            },
            {
                "name": "RSVP Reminder",
                "category": "reminder",
                "content": (
                    "Hi {{guest_name}}, friendly reminder to RSVP for {{couple_names}}'s wedding by {{rsvp_deadline}}! "
                    "RSVP here: {{chat_link}}"
                ),
            },
            {
                "name": "One Week Countdown",
                "category": "reminder",
                "content": (
                    "{{guest_name}}, just ONE WEEK until {{couple_names}}'s wedding! "
                    "Date: {{wedding_date}}. Dress code: {{dress_code}}. "
                    "Questions? {{chat_link}}"
                ),
            },
            {
                "name": "Day Before Reminder",
                "category": "reminder",
                "content": (
                    "{{guest_name}}, the big day is TOMORROW! "
                    "{{couple_names}}'s ceremony starts at {{wedding_time}}. "
                    "Venue: {{ceremony_venue}}. See you there!"
                ),
            },
            {
                "name": "Day Of Logistics",
                "category": "update",
                "content": (
                    "Good morning {{guest_name}}! Today's the day! "
                    "Ceremony at {{wedding_time}} at {{ceremony_venue}}. "
                    "Dress code: {{dress_code}}. Can't wait to celebrate!"
                ),
            },
            {
                "name": "Thank You",
                "category": "update",
                "content": (
                    "Thank you for celebrating with us, {{guest_name}}! "
                    "Your presence made our day extra special. - {{partner1}} & {{partner2}}"
                ),
            },
        ]

        for tmpl_data in default_templates:
            template = SMSTemplate(
                wedding_id=wedding_id,
                name=tmpl_data["name"],
                content=tmpl_data["content"],
                category=tmpl_data["category"],
                is_default=False  # Not a system default, so editable
            )
            db.add(template)

        await db.commit()

        # Reload templates after seeding
        result = await db.execute(
            select(SMSTemplate).where(SMSTemplate.wedding_id == wedding_id)
        )
        templates = list(result.scalars().all())

    return templates


@router.get("/templates/variables", response_model=List[dict])
async def get_template_variables():
    """Get list of available template variables."""
    return template_service.get_available_variables()


@router.post("/{wedding_id}/templates", response_model=dict)
async def create_template(
    wedding_id: str,
    data: TemplateCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a custom SMS template."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Validate template
    is_valid, unknown_vars = template_service.validate_template(data.content)
    if not is_valid:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown template variables: {', '.join(unknown_vars)}"
        )

    template = SMSTemplate(
        wedding_id=wedding_id,
        name=data.name,
        content=data.content,
        category=data.category,
        is_default=False
    )
    db.add(template)
    await db.commit()
    await db.refresh(template)

    return {"id": template.id, "message": "Template created"}


@router.patch("/{wedding_id}/templates/{template_id}", response_model=dict)
async def update_template(
    wedding_id: str,
    template_id: str,
    data: TemplateUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a template."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(SMSTemplate).where(
            SMSTemplate.id == template_id,
            SMSTemplate.wedding_id == wedding_id
        )
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    update_data = data.model_dump(exclude_unset=True)

    # Validate content if being updated
    if "content" in update_data:
        is_valid, unknown_vars = template_service.validate_template(update_data["content"])
        if not is_valid:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown template variables: {', '.join(unknown_vars)}"
            )

    for field, value in update_data.items():
        setattr(template, field, value)

    await db.commit()
    return {"id": template_id, "message": "Template updated"}


@router.delete("/{wedding_id}/templates/{template_id}", response_model=dict)
async def delete_template(
    wedding_id: str,
    template_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a template."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(SMSTemplate).where(
            SMSTemplate.id == template_id,
            SMSTemplate.wedding_id == wedding_id
        )
    )
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    await db.delete(template)
    await db.commit()
    return {"message": "Template deleted"}


# --- SMS Campaign Endpoints ---

@router.post("/{wedding_id}/sms/send", response_model=dict)
async def send_sms_blast(
    wedding_id: str,
    data: SMSBlastRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Send SMS blast immediately to selected recipients."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if not twilio_service.is_configured:
        raise HTTPException(status_code=503, detail="SMS service not configured")

    # Get wedding for template rendering
    result = await db.execute(select(Wedding).where(Wedding.id == wedding_id))
    wedding = result.scalar_one_or_none()
    if not wedding:
        raise HTTPException(status_code=404, detail="Wedding not found")

    # Get eligible recipients
    recipients = await get_eligible_recipients(
        db, wedding_id, data.recipient_type, data.recipient_filter
    )

    if not recipients:
        raise HTTPException(status_code=400, detail="No eligible recipients found")

    # Create scheduled message record
    scheduled = ScheduledMessage(
        wedding_id=wedding_id,
        name=f"Blast {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
        message_content=data.message,
        recipient_type=data.recipient_type,
        recipient_filter=data.recipient_filter,
        schedule_type="fixed",
        scheduled_at=datetime.utcnow(),
        status="sending",
        total_recipients=len(recipients)
    )
    db.add(scheduled)
    await db.commit()
    await db.refresh(scheduled)

    # Send to each recipient
    sent = 0
    failed = 0

    for guest in recipients:
        # Render template with guest/wedding data
        rendered_message = template_service.render(data.message, guest, wedding)

        # Send SMS
        result = await twilio_service.send_sms(guest.phone_number, rendered_message)

        # Log the send
        log = MessageLog(
            wedding_id=wedding_id,
            guest_id=guest.id,
            scheduled_message_id=scheduled.id,
            phone_number=guest.phone_number,
            message_content=rendered_message,
            twilio_sid=result.get("sid"),
            status=result.get("status", "failed"),
            error_code=result.get("error_code"),
            error_message=result.get("error_message"),
            sent_at=datetime.utcnow() if result.get("success") else None
        )
        db.add(log)

        if result.get("success"):
            sent += 1
        else:
            failed += 1

    # Update scheduled message status
    scheduled.sent_count = sent
    scheduled.failed_count = failed
    scheduled.status = "sent" if failed == 0 else "partially_sent"
    scheduled.sent_at = datetime.utcnow()
    await db.commit()

    return {
        "message": f"Sent to {sent} of {len(recipients)} recipients",
        "sent": sent,
        "failed": failed,
        "scheduled_message_id": scheduled.id
    }


@router.post("/{wedding_id}/sms/schedule", response_model=dict)
async def schedule_message(
    wedding_id: str,
    data: ScheduleMessageRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Schedule a message for future delivery."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Validate schedule
    if data.schedule_type == "fixed" and not data.scheduled_at:
        raise HTTPException(status_code=400, detail="scheduled_at required for fixed schedule")

    if data.schedule_type == "relative":
        if not data.relative_to or data.relative_days is None:
            raise HTTPException(
                status_code=400,
                detail="relative_to and relative_days required for relative schedule"
            )

    # Get recipient count
    recipients = await get_eligible_recipients(
        db, wedding_id, data.recipient_type, data.recipient_filter
    )

    scheduled = ScheduledMessage(
        wedding_id=wedding_id,
        name=data.name,
        message_content=data.message,
        recipient_type=data.recipient_type,
        recipient_filter=data.recipient_filter,
        schedule_type=data.schedule_type,
        scheduled_at=data.scheduled_at,
        relative_to=data.relative_to,
        relative_days=data.relative_days,
        status="scheduled",
        total_recipients=len(recipients)
    )
    db.add(scheduled)
    await db.commit()
    await db.refresh(scheduled)

    return {"id": scheduled.id, "message": "Message scheduled"}


@router.get("/{wedding_id}/sms/scheduled", response_model=List[ScheduledMessageResponse])
async def list_scheduled_messages(
    wedding_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all scheduled messages for a wedding."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(ScheduledMessage)
        .where(ScheduledMessage.wedding_id == wedding_id)
        .order_by(ScheduledMessage.created_at.desc())
    )
    return result.scalars().all()


@router.delete("/{wedding_id}/sms/scheduled/{message_id}", response_model=dict)
async def cancel_scheduled_message(
    wedding_id: str,
    message_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Cancel a scheduled message."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(ScheduledMessage).where(
            ScheduledMessage.id == message_id,
            ScheduledMessage.wedding_id == wedding_id
        )
    )
    scheduled = result.scalar_one_or_none()
    if not scheduled:
        raise HTTPException(status_code=404, detail="Scheduled message not found")

    if scheduled.status in ("sent", "partially_sent"):
        raise HTTPException(status_code=400, detail="Cannot cancel already sent message")

    scheduled.status = "cancelled"
    await db.commit()
    return {"message": "Scheduled message cancelled"}


@router.get("/{wedding_id}/sms/history", response_model=List[MessageLogResponse])
async def get_sms_history(
    wedding_id: str,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get SMS send history for a wedding."""
    if current_user.wedding_id != wedding_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(
        select(MessageLog)
        .where(MessageLog.wedding_id == wedding_id)
        .order_by(MessageLog.created_at.desc())
        .limit(limit)
    )
    return result.scalars().all()


# --- Webhook Endpoints (for Twilio callbacks) ---

async def verify_twilio_signature(
    request: Request,
    x_twilio_signature: Optional[str] = Header(None, alias="X-Twilio-Signature")
) -> bool:
    """Verify Twilio webhook request signature."""
    # In production, signature is required
    if IS_PRODUCTION and not x_twilio_signature:
        logger.warning("Twilio webhook received without signature")
        raise HTTPException(status_code=403, detail="Missing Twilio signature")

    # If no signature provided (dev mode), skip validation
    if not x_twilio_signature:
        return True

    # Get the full URL of the request
    url = str(request.url)

    # Get form parameters
    form_data = await request.form()
    params = {key: str(value) for key, value in form_data.items()}

    # Validate the signature
    if not twilio_service.validate_webhook_request(url, params, x_twilio_signature):
        logger.warning(f"Invalid Twilio signature for webhook: {url}")
        raise HTTPException(status_code=403, detail="Invalid Twilio signature")

    return True


@router.post("/webhooks/status")
async def sms_status_webhook(
    request: Request,
    MessageSid: str = Form(...),
    MessageStatus: str = Form(...),
    To: str = Form(None),
    ErrorCode: str = Form(None),
    ErrorMessage: str = Form(None),
    x_twilio_signature: Optional[str] = Header(None, alias="X-Twilio-Signature"),
    db: AsyncSession = Depends(get_db)
):
    """Handle Twilio delivery status callbacks."""
    # Verify Twilio signature
    await verify_twilio_signature(request, x_twilio_signature)

    # Find the message log by Twilio SID
    result = await db.execute(
        select(MessageLog).where(MessageLog.twilio_sid == MessageSid)
    )
    log = result.scalar_one_or_none()

    if log:
        log.status = MessageStatus
        if MessageStatus == "delivered":
            log.delivered_at = datetime.utcnow()
        if ErrorCode:
            log.error_code = ErrorCode
            log.error_message = ErrorMessage
        await db.commit()

    return {"status": "ok"}


@router.post("/webhooks/inbound")
async def sms_inbound_webhook(
    request: Request,
    From: str = Form(...),
    Body: str = Form(...),
    x_twilio_signature: Optional[str] = Header(None, alias="X-Twilio-Signature"),
    db: AsyncSession = Depends(get_db)
):
    """Handle inbound SMS (opt-out requests)."""
    # Verify Twilio signature
    await verify_twilio_signature(request, x_twilio_signature)

    body_upper = Body.upper().strip()

    # Check for opt-out keywords
    opt_out_keywords = ["STOP", "STOPALL", "UNSUBSCRIBE", "CANCEL", "END", "QUIT"]

    if any(keyword in body_upper for keyword in opt_out_keywords):
        # Find all guests with this phone number and opt them out
        result = await db.execute(
            select(Guest).where(Guest.phone_number == From)
        )
        guests = result.scalars().all()

        for guest in guests:
            guest.opted_out = True
            guest.opted_out_at = datetime.utcnow()

        await db.commit()

        # Note: Twilio automatically sends "You have been unsubscribed" for STOP messages
        # when using a Messaging Service with Advanced Opt-Out enabled

    return {"status": "ok"}
